#Disclaimer: man stands for manipulation, and not, I'll not argue with feminists..
import time
import os
import DataSource as d
from openpyxl import load_workbook

#logs message and shows whos called.
def log(message):
    import inspect
    print(time.ctime(),' [',inspect.stack()[1][3],']: ', message)

#converts xls to xlsx using win32com library and excludes the xls files after successful conversion
def convert_xls_xlsx(file):
    import win32com.client as win32 
    xl = win32.gencache.EnsureDispatch('Excel.Application')
    wb = xl.Workbooks.Open(file)
    wb.SaveAs(file+'x', FileFormat=51) #51 for xlsx/56 for xls
    wb.Close()
    xl.Application.Quit()
    log(file+' has been converted to '+file+'x')

#opens file in Excel for visualization
def open_in_excel(filepath):
    from win32com.client import Dispatch
    log('Opening in ExceL: '+ filepath)
    wb=Dispatch("Excel.Application")
    wb.Visible=True
    wb.Workbooks.Open(filepath)

#fills cells with no data.
#wb is workbook, sh is sheet, r is row, and c is column
def purge_data(sh, r_ini, r_fin, c_ini, c_fin):
    for r in range (r_ini,r_fin):
        for c in range (c_ini, c_fin):
            sh.cell(column=c, row=r).value = ""
    log(str(sh)+' has been purged!') 

#it does what it seems to do: fetch data.
def fetch_data(source, destination, r_ini, r_fin, c_ini, c_fin):
    for r in range(r_ini, r_fin):
        for c in range(c_ini, c_fin):
            destination.cell(row=r, column=c).value = source.cell(row=r, column=c).value
    log(destination.title+ ' has been graciously fed by '+source.title)
    
#it cleans the already paid bills
def clean_data(ws, r_ini):
    for r in range (r_ini, ws.max_row):
        s = ws.cell(row=r, column=1).value
        if(s is not None and s.find('PAGO')!=-1):#check if its None and if it is paid.
            ws.cell(row=r,column=15).value = float(0.00)
            log(ws.title+ ': Row '+str(r)+' current value has been erased. Reason: Already paid.')

#format as float and add currency symbol in the output values
def format_currency_data(ws, c, r_ini, r_fin):
    for r in range(r_ini, r_fin):
        _cell = ws.cell(column = c, row = r)
        if(_cell.value is not None and _cell.value != ''):
           _cell.value = float(str(_cell.value).replace('.','').replace(',','.'))
           _cell.number_format ='#,##0.00R$'

def send_mail(mailing_list, file_list, message):
    log('NOT IMPLEMENTED YET')

def backup(filetype):
    if(filetype=='xls'):
        os.system('cmd /c move /Y '+d.VE09+' .\\backup')
        os.system('cmd /c move /Y '+d.AV09+' .\\backup')
        os.system('cmd /c move /Y '+d.VE28+' .\\backup')
        os.system('cmd /c move /Y '+d.AV28+' .\\backup')
        log('.xls files has been backed up')
    if(filetype=='xlsx'):
        os.system('cmd /c move /Y '+d.VE09X+' .\\backup')
        os.system('cmd /c move /Y '+d.AV09X+' .\\backup')
        os.system('cmd /c move /Y '+d.VE28X+' .\\backup')
        os.system('cmd /c move /Y '+d.AV28X+' .\\backup')
        log('.xlsx files has been backed up')




    

      