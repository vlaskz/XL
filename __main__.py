from openpyxl import load_workbook
import Manipulation as m, Const_Bradesco as b, DataSource as d #some people certainly will brag about this.

#PART ONE - Get the data
#move older result(xlsx) files to backup folder
m.backup('xlsx')

#converts xls worksheets to xlsx
m.log('starting data workbook conversion')
m.convert_xls_xlsx(d.VE09)
m.convert_xls_xlsx(d.AV09)
m.convert_xls_xlsx(d.VE28)
m.convert_xls_xlsx(d.AV28)
m.log('finishing workbook conversion')

#loads the new datafiles into memory - ok
m.log('starting workbook data loading')
VE09X = load_workbook(d.VE09X)
AV09X = load_workbook(d.AV09X)
VE28X = load_workbook(d.VE28X)
AV28X = load_workbook(d.AV28X)
RECBX = load_workbook(d.RECBX, read_only=False)
m.log('finishing workbook data loading')

#clears the old data from the worksheets.
m.log('starting cleaning data from main workbook')
m.purge_data(RECBX['VE09BRAD'],b.R_INI,VE09X.active.max_row,b.C_INI,b.C_FIN)
m.purge_data(RECBX['AV09BRAD'],b.R_INI,AV09X.active.max_row,b.C_INI,b.C_FIN)
m.purge_data(RECBX['VE28BRAD'],b.R_INI,VE28X.active.max_row,b.C_INI,b.C_FIN)
m.purge_data(RECBX['AV28BRAD'],b.R_INI,AV28X.active.max_row,b.C_INI,b.C_FIN)
m.log('finishing cleaning data from main workbook')

#copy data from datafiles to main workbook
m.log('starting copying data from data workbooks to main workbook')
m.fetch_data(VE09X.active,RECBX['VE09BRAD'],b.R_INI,VE09X.active.max_row,b.C_INI,b.C_FIN)
m.fetch_data(AV09X.active,RECBX['AV09BRAD'],b.R_INI,AV09X.active.max_row,b.C_INI,b.C_FIN)
m.fetch_data(VE28X.active,RECBX['VE28BRAD'],b.R_INI,VE28X.active.max_row,b.C_INI,b.C_FIN)
m.fetch_data(AV28X.active,RECBX['AV28BRAD'],b.R_INI,AV28X.active.max_row,b.C_INI,b.C_FIN)
m.log('finishing copying data from data workbooks to main workbook')

#PART 2: let's unbloat and unjunk the data (yes I'm a neological person)
m.log('starting unbloating data')
m.clean_data(RECBX['VE09BRAD'],b.R_INI)
m.clean_data(RECBX['AV09BRAD'],b.R_INI)
m.clean_data(RECBX['VE28BRAD'],b.R_INI)
m.clean_data(RECBX['AV28BRAD'],b.R_INI)
m.log('finishing unbloating data')

#format data in the "O" column as float and set the field as currency (at the moment is BRL(R$))
m.log('starting currency data formatting in main workbook')
m.format_currency_data(RECBX['VE09BRAD'],b.CUR_COL,b.R_INI,RECBX['VE09BRAD'].max_row)
m.format_currency_data(RECBX['AV09BRAD'],b.CUR_COL,b.R_INI,RECBX['AV09BRAD'].max_row)
m.format_currency_data(RECBX['VE28BRAD'],b.CUR_COL,b.R_INI,RECBX['VE28BRAD'].max_row)
m.format_currency_data(RECBX['AV28BRAD'],b.CUR_COL,b.R_INI,RECBX['AV28BRAD'].max_row)
m.log('finishing currency data formatting in main workbook')

m.log('saving main workbook')
RECBX.save(d.RECBX)
m.backup('xls')
m.backup('xlsx')

m.log('opening main workbook')
m.open_in_excel(d.RECBX)

#Part 3: Send RECBX to email.

#sends emails to mail_listing
m.log('starting sending main workbook to mailing list')
m.send_mail('list','annex', 'msg')